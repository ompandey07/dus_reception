from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from datetime import datetime, date, timedelta
from authapp.decorators import login_required_dual
from authapp.models import CustomUser
from .models import Booking, ActivityLog
import json


# ============================================================
# BOOKING REPORTS VIEW
# ============================================================
@login_required_dual(login_url='/unauthorized/')
def booking_reports_view(request):
    """Render the booking reports page"""
    custom_users = CustomUser.objects.all()
    
    # Get current date info
    today = date.today()
    
    # Get event types for filter
    event_types = Booking.objects.values_list('event_type', flat=True).distinct()
    
    # Get time slots for filter
    time_slots = Booking.TIME_SLOT_CHOICES
    
    context = {
        'custom_users': custom_users,
        'event_types': event_types,
        'time_slots': time_slots,
        'today': today
    }
    return render(request, 'admin/booking_reports.html', context)


@login_required_dual(login_url='/unauthorized/')
@require_http_methods(["GET"])
def get_booking_reports(request):
    """API endpoint to get filtered booking reports with pagination"""
    try:
        # Get filter parameters
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        date_from = request.GET.get('date_from', None)
        date_to = request.GET.get('date_to', None)
        event_type = request.GET.get('event_type', None)
        time_slot = request.GET.get('time_slot', None)
        created_by_filter = request.GET.get('created_by', None)
        search = request.GET.get('search', None)
        min_advance = request.GET.get('min_advance', None)
        max_advance = request.GET.get('max_advance', None)
        
        # Start with all bookings
        bookings = Booking.objects.all()
        
        # Apply filters
        if date_from:
            bookings = bookings.filter(booking_date__gte=date_from)
        if date_to:
            bookings = bookings.filter(booking_date__lte=date_to)
        if event_type:
            bookings = bookings.filter(event_type=event_type)
        if time_slot:
            bookings = bookings.filter(time_slot=time_slot)
        if min_advance:
            bookings = bookings.filter(advance_given__gte=min_advance)
        if max_advance:
            bookings = bookings.filter(advance_given__lte=max_advance)
        
        # Apply created_by filter
        if created_by_filter:
            if created_by_filter.startswith('user_'):
                user_id = created_by_filter.replace('user_', '')
                bookings = bookings.filter(created_by_user_id=user_id)
            elif created_by_filter.startswith('custom_'):
                custom_id = created_by_filter.replace('custom_', '')
                bookings = bookings.filter(created_by_custom_id=custom_id)
        
        # Apply search
        if search:
            bookings = bookings.filter(
                Q(client_name__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(email__icontains=search) |
                Q(event_type__icontains=search) |
                Q(menu_type__icontains=search)
            )
        
        # Order by booking date descending
        bookings = bookings.order_by('-booking_date', 'time_slot')
        
        # Calculate statistics before pagination
        total_bookings = bookings.count()
        total_advance = bookings.aggregate(Sum('advance_given'))['advance_given__sum'] or 0
        
        # Get event type breakdown
        event_breakdown = bookings.values('event_type').annotate(
            count=Count('id'),
            total_advance=Sum('advance_given')
        ).order_by('-count')
        
        # Paginate
        paginator = Paginator(bookings, per_page)
        page_obj = paginator.get_page(page)
        
        # Convert bookings to JSON
        bookings_data = []
        for booking in page_obj:
            from .views import get_nepali_date
            nepali_date = get_nepali_date(booking.booking_date)
            
            # Get time slot display
            time_slot_display = dict(Booking.TIME_SLOT_CHOICES).get(booking.time_slot, booking.time_slot)
            
            bookings_data.append({
                'id': booking.id,
                'client_name': booking.client_name,
                'booking_date': booking.booking_date.strftime('%Y-%m-%d'),
                'booking_date_formatted': booking.booking_date.strftime('%B %d, %Y'),
                'booking_date_nepali': nepali_date['formatted_nepali'] if nepali_date else '',
                'time_slot': booking.time_slot,
                'time_slot_display': time_slot_display,
                'phone_number': booking.phone_number,
                'email': booking.email or '',
                'event_type': booking.event_type,
                'menu_type': booking.menu_type or '',
                'no_of_pax': booking.no_of_pax or '',
                'advance_given': float(booking.advance_given),
                'created_by': booking.get_creator_name(),
                'created_at': booking.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Log report generation activity
        from .views import log_activity, get_client_ip
        performed_by_user = None
        performed_by_custom = None
        
        if request.user.is_authenticated:
            performed_by_user = request.user
        else:
            custom_user_id = request.COOKIES.get("custom_user_id")
            if custom_user_id:
                try:
                    performed_by_custom = CustomUser.objects.get(id=custom_user_id)
                except CustomUser.DoesNotExist:
                    pass
        
        filter_desc = []
        if date_from: filter_desc.append(f"from {date_from}")
        if date_to: filter_desc.append(f"to {date_to}")
        if event_type: filter_desc.append(f"event: {event_type}")
        
        log_activity(
            'view',
            'booking',
            description=f'Generated booking report ({total_bookings} bookings{", " + ", ".join(filter_desc) if filter_desc else ""})',
            request=request,
            performed_by_user=performed_by_user,
            performed_by_custom=performed_by_custom
        )
        
        return JsonResponse({
            'bookings': bookings_data,
            'statistics': {
                'total_bookings': total_bookings,
                'total_advance': float(total_advance),
                'event_breakdown': list(event_breakdown)
            },
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': total_bookings,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'per_page': per_page
            }
        }, status=200)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required_dual(login_url='/unauthorized/')
@require_http_methods(["GET"])
def export_booking_reports(request):
    """Export booking reports to Excel with formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        return export_booking_reports_csv(request)
    
    try:
        # Get same filters as report view
        date_from = request.GET.get('date_from', None)
        date_to = request.GET.get('date_to', None)
        event_type = request.GET.get('event_type', None)
        time_slot = request.GET.get('time_slot', None)
        created_by_filter = request.GET.get('created_by', None)
        search = request.GET.get('search', None)
        
        bookings = Booking.objects.all()
        
        # Apply same filters
        if date_from:
            bookings = bookings.filter(booking_date__gte=date_from)
        if date_to:
            bookings = bookings.filter(booking_date__lte=date_to)
        if event_type:
            bookings = bookings.filter(event_type=event_type)
        if time_slot:
            bookings = bookings.filter(time_slot=time_slot)
        if created_by_filter:
            if created_by_filter.startswith('user_'):
                user_id = created_by_filter.replace('user_', '')
                bookings = bookings.filter(created_by_user_id=user_id)
            elif created_by_filter.startswith('custom_'):
                custom_id = created_by_filter.replace('custom_', '')
                bookings = bookings.filter(created_by_custom_id=custom_id)
        if search:
            bookings = bookings.filter(
                Q(client_name__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(email__icontains=search) |
                Q(event_type__icontains=search) |
                Q(menu_type__icontains=search)
            )
        
        bookings = bookings.order_by('-booking_date', 'time_slot')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Booking Reports"
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Define headers
        headers = [
            'Client Name', 'Booking Date', 'Time Slot', 
            'Phone Number', 'Email', 'Event Type', 'Menu Type', 
            'No. of Pax', 'Advance Given', 'Created By', 'Created At'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Write data
        for row_num, booking in enumerate(bookings, 2):
            time_slot_display = dict(Booking.TIME_SLOT_CHOICES).get(booking.time_slot, booking.time_slot)
            
            data = [
                booking.client_name,
                booking.booking_date.strftime('%Y-%m-%d'),
                time_slot_display,
                booking.phone_number,
                booking.email or '',
                booking.event_type,
                booking.menu_type or '',
                booking.no_of_pax or '',
                float(booking.advance_given),
                booking.get_creator_name(),
                booking.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                
                # Apply alignment based on column
                if col_num in [2, 3, 8, 9]:  # Date, time slot, pax, advance - center
                    cell.alignment = center_alignment
                else:
                    cell.alignment = cell_alignment
                
                # Format advance amount
                if col_num == 9:  # Advance Given column
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        column_widths = {
            1: 20,  # Client Name
            2: 12,  # Booking Date
            3: 15,  # Time Slot
            4: 15,  # Phone Number
            5: 25,  # Email
            6: 15,  # Event Type
            7: 20,  # Menu Type
            8: 12,  # No. of Pax
            9: 12,  # Advance Given
            10: 18, # Created By
            11: 18  # Created At
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="booking_report_{date.today()}.xlsx"'
        
        wb.save(response)
        
        # Log export activity
        from .views import log_activity
        performed_by_user = None
        performed_by_custom = None
        
        if request.user.is_authenticated:
            performed_by_user = request.user
        else:
            custom_user_id = request.COOKIES.get("custom_user_id")
            if custom_user_id:
                try:
                    performed_by_custom = CustomUser.objects.get(id=custom_user_id)
                except CustomUser.DoesNotExist:
                    pass
        
        log_activity(
            'export',
            'booking',
            description=f'Exported {bookings.count()} bookings to Excel',
            request=request,
            performed_by_user=performed_by_user,
            performed_by_custom=performed_by_custom
        )
        
        return response
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def export_booking_reports_csv(request):
    """Fallback CSV export if openpyxl is not available"""
    import csv
    
    try:
        # Get same filters as report view
        date_from = request.GET.get('date_from', None)
        date_to = request.GET.get('date_to', None)
        event_type = request.GET.get('event_type', None)
        time_slot = request.GET.get('time_slot', None)
        created_by_filter = request.GET.get('created_by', None)
        search = request.GET.get('search', None)
        
        bookings = Booking.objects.all()
        
        # Apply same filters
        if date_from:
            bookings = bookings.filter(booking_date__gte=date_from)
        if date_to:
            bookings = bookings.filter(booking_date__lte=date_to)
        if event_type:
            bookings = bookings.filter(event_type=event_type)
        if time_slot:
            bookings = bookings.filter(time_slot=time_slot)
        if created_by_filter:
            if created_by_filter.startswith('user_'):
                user_id = created_by_filter.replace('user_', '')
                bookings = bookings.filter(created_by_user_id=user_id)
            elif created_by_filter.startswith('custom_'):
                custom_id = created_by_filter.replace('custom_', '')
                bookings = bookings.filter(created_by_custom_id=custom_id)
        if search:
            bookings = bookings.filter(
                Q(client_name__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(email__icontains=search) |
                Q(event_type__icontains=search) |
                Q(menu_type__icontains=search)
            )
        
        bookings = bookings.order_by('-booking_date', 'time_slot')
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="booking_report_{date.today()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Client Name', 'Booking Date', 'Time Slot', 
                        'Phone Number', 'Email', 'Event Type', 'Menu Type', 
                        'No. of Pax', 'Advance Given', 'Created By', 'Created At'])
        
        for booking in bookings:
            time_slot_display = dict(Booking.TIME_SLOT_CHOICES).get(booking.time_slot, booking.time_slot)
            writer.writerow([
                booking.client_name,
                booking.booking_date.strftime('%Y-%m-%d'),
                time_slot_display,
                booking.phone_number,
                booking.email or '',
                booking.event_type,
                booking.menu_type or '',
                booking.no_of_pax or '',
                float(booking.advance_given),
                booking.get_creator_name(),
                booking.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)